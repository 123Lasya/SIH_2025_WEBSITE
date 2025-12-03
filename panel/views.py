from django.shortcuts import render
from django.contrib.auth.decorators import user_passes_test

from accounts.models import FarmerProfile, BuyerBusinessProfile, BuyerTraderProfile
from farmers.models import FarmerContract, FarmerHedge
from buyers.models import BuyerHedge
from farmers.models import Alert

from django.db.models import Count
from django.db.models.functions import ExtractMonth
import json


def admin_only(view_func):
    return user_passes_test(lambda u: u.is_staff or u.is_superuser)(view_func)


# -------------------------------
# ADMIN DASHBOARD
# -------------------------------
@admin_only
def admin_dashboard(request):

    total_farmers = FarmerProfile.objects.count()
    total_buyers = BuyerBusinessProfile.objects.count() + BuyerTraderProfile.objects.count()
    total_contracts = FarmerContract.objects.count()
    total_hedges = FarmerHedge.objects.count() + BuyerHedge.objects.count()

    # Hedges per month
    hedge_months = (
        FarmerHedge.objects.annotate(month=ExtractMonth("created_at"))
        .values("month")
        .annotate(count=Count("id"))
        .order_by("month")
    )

    buyer_hedge_months = (
        BuyerHedge.objects.annotate(month=ExtractMonth("created_at"))
        .values("month")
        .annotate(count=Count("id"))
        .order_by("month")
    )

    hedge_chart = [0] * 12
    for h in hedge_months:
        hedge_chart[h["month"] - 1] += h["count"]

    for h in buyer_hedge_months:
        hedge_chart[h["month"] - 1] += h["count"]

    # Contracts per month
    contract_months = (
        FarmerContract.objects.annotate(month=ExtractMonth("created_at"))
        .values("month")
        .annotate(count=Count("id"))
        .order_by("month")
    )

    contract_chart = [0] * 12
    for c in contract_months:
        contract_chart[c["month"] - 1] = c["count"]

    context = {
        "farmers": total_farmers,
        "buyers": total_buyers,
        "contracts": total_contracts,
        "hedges": total_hedges,
        "hedge_chart": json.dumps(hedge_chart),
        "contract_chart": json.dumps(contract_chart),
    }

    return render(request, "panel/dashboard.html", context)


# --------------------------------
# LIST PAGES
# --------------------------------

@admin_only
def admin_farmers(request):
    farmers = FarmerProfile.objects.all()
    return render(request, "panel/farmers_list.html", {"farmers": farmers})


@admin_only
def admin_buyers(request):
    business = BuyerBusinessProfile.objects.all()
    traders = BuyerTraderProfile.objects.all()
    return render(request, "panel/buyers_list.html", {"business": business, "traders": traders})


@admin_only
def admin_contracts(request):
    contracts = FarmerContract.objects.all().order_by("-id")
    return render(request, "panel/contracts_list.html", {"contracts": contracts})


@admin_only
def admin_hedges(request):
    farmer_hedges = FarmerHedge.objects.all()
    buyer_hedges = BuyerHedge.objects.all()
    return render(
        request,
        "panel/hedges_list.html",
        {"farmer_hedges": farmer_hedges, "buyer_hedges": buyer_hedges},
    )


@admin_only
def admin_alerts(request):
    alerts = Alert.objects.all().order_by("-id")
    return render(request, "panel/alerts_list.html", {"alerts": alerts})

import csv
from django.http import HttpResponse
from openpyxl import Workbook

@admin_only
def export_farmers_csv(request):
    farmers = FarmerProfile.objects.all()

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="farmers.csv"'

    writer = csv.writer(response)
    writer.writerow(["ID", "Name", "Email", "District"])

    for f in farmers:
        writer.writerow([
            f.id,
            f.user.first_name,
            f.user.email,
            f.district
        ])

    return response
@admin_only
def export_farmers_excel(request):
    farmers = FarmerProfile.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Farmers"

    ws.append(["ID", "Name", "Email", "District"])

    for f in farmers:
        ws.append([f.id, f.user.first_name, f.user.email, f.district])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="farmers.xlsx"'
    wb.save(response)
    return response
@admin_only
def export_buyers_csv(request):
    business = BuyerBusinessProfile.objects.all()
    traders = BuyerTraderProfile.objects.all()

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="buyers.csv"'

    writer = csv.writer(response)
    writer.writerow(["Type", "Name", "Email", "District/Factory"])

    for b in business:
        writer.writerow(["Business", b.user.first_name, b.business_email, b.factory_address])

    for t in traders:
        writer.writerow(["Trader", t.user.first_name, t.user.email, t.district])

    return response
@admin_only
def export_buyers_excel(request):
    business = BuyerBusinessProfile.objects.all()
    traders = BuyerTraderProfile.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Buyers"

    ws.append(["Type", "Name", "Email", "District/Factory"])

    for b in business:
        ws.append(["Business", b.user.first_name, b.business_email, b.factory_address])

    for t in traders:
        ws.append(["Trader", t.user.first_name, t.user.email, t.district])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="buyers.xlsx"'
    wb.save(response)
    return response
@admin_only
def export_contracts_csv(request):
    contracts = FarmerContract.objects.all()

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="contracts.csv"'

    writer = csv.writer(response)
    writer.writerow(["ID", "Farmer", "Crop", "Qty", "Status", "Start", "End"])

    for c in contracts:
        writer.writerow([
            c.id,
            c.farmer.user.first_name,
            c.crop,
            c.quantity,
            c.buyer_status,
            c.start_date,
            c.end_date
        ])

    return response
@admin_only
def export_contracts_excel(request):
    contracts = FarmerContract.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Contracts"

    ws.append(["ID", "Farmer", "Crop", "Qty", "Status", "Start", "End"])

    for c in contracts:
        ws.append([
            c.id,
            c.farmer.user.first_name,
            c.crop,
            c.quantity,
            c.buyer_status,
            c.start_date,
            c.end_date
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="contracts.xlsx"'
    wb.save(response)
    return response

@admin_only
def export_hedges_csv(request):
    farmer_hedges = FarmerHedge.objects.all()
    buyer_hedges = BuyerHedge.objects.all()

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="hedges.csv"'

    writer = csv.writer(response)
    writer.writerow(["Type", "User", "Crop", "Qty", "Price", "End Date"])

    for h in farmer_hedges:
        writer.writerow(["Farmer", h.farmer.user.first_name, h.crop, h.quantity, h.hedge_price, h.end_date])

    for h in buyer_hedges:
        writer.writerow(["Buyer", h.buyer.email, h.crop, h.quantity, h.hedge_price, h.end_date])

    return response
@admin_only
def export_hedges_excel(request):
    farmer_hedges = FarmerHedge.objects.all()
    buyer_hedges = BuyerHedge.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Hedges"

    ws.append(["Type", "User", "Crop", "Qty", "Price", "End Date"])

    for h in farmer_hedges:
        ws.append(["Farmer", h.farmer.user.first_name, h.crop, h.quantity, h.hedge_price, h.end_date])

    for h in buyer_hedges:
        ws.append(["Buyer", h.buyer.email, h.crop, h.quantity, h.hedge_price, h.end_date])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="hedges.xlsx"'
    wb.save(response)
    return response
